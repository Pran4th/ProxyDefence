"""
Phase 1: Study Every Source — Deep schema inference from actual data.
Handles CSV (with/without headers), Excel (.xlsx), ZIP archives, and text files.
"""
import csv
import json
import os
import sys
import zipfile
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
from typing import Any

RESEARCH_DIR = Path(__file__).parent
DATASETS_DIR = RESEARCH_DIR.parent / "datasets"
OUTPUT_DIR = RESEARCH_DIR / "inventory"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(RESEARCH_DIR.parent / "services/ml-platform"))
sys.path.insert(0, str(RESEARCH_DIR.parent))


def fmt_bytes(n):
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


def infer_dtype(values, max_sample=200):
    types = Counter()
    for v in values:
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        v_str = str(v).strip()
        if not v_str:
            continue
        try:
            float(v_str)
            types["numeric"] += 1
            continue
        except ValueError:
            pass
        if v_str.lower() in ("true", "false", "t", "f", "yes", "no"):
            types["boolean"] += 1
            continue
        types["string"] += 1
    if not types:
        return "unknown"
    if types.get("numeric", 0) > 0.8 * sum(types.values()):
        return "numeric"
    return "string"


def detect_has_header(path, sample_rows=20):
    """Detect if CSV has a header row by checking if first row values look like column names."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            first_row = next(reader, [])
            second_row = next(reader, [])
            
            if not first_row or not second_row:
                return False
            
            # Check if first row values look like column names (strings, not numbers)
            numeric_count = 0
            for v in first_row:
                try:
                    float(v.strip())
                    numeric_count += 1
                except ValueError:
                    pass
            
            # If more than half of first row values are numeric, it's likely data
            if numeric_count > len(first_row) / 2:
                return False
            
            # Check if second row has similar number of columns
            if len(second_row) != len(first_row):
                return False
            
            return True
    except Exception:
        return False


def analyze_csv(path, max_rows=50000):
    try:
        has_header = detect_has_header(path)
        
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            headers = None
            col_data = None
            row_count = 0
            
            for i, row in enumerate(reader):
                if not row or all(c.strip() == "" for c in row):
                    continue
                    
                if headers is None:
                    if has_header:
                        headers = [c.strip() for c in row]
                    else:
                        headers = [f"col_{j}" for j in range(len(row))]
                    col_data = {h: {"nulls": 0, "total": 0, "values": []} for h in headers}
                    if not has_header:
                        # Process first row as data
                        for j, h in enumerate(headers):
                            v = row[j] if j < len(row) else ""
                            col_data[h]["total"] += 1
                            if v == "" or v is None:
                                col_data[h]["nulls"] += 1
                            if len(col_data[h]["values"]) < 200:
                                col_data[h]["values"].append(v)
                        row_count += 1
                    continue
                
                row_count += 1
                if row_count > max_rows:
                    break
                
                for j, h in enumerate(headers):
                    v = row[j] if j < len(row) else ""
                    col_data[h]["total"] += 1
                    if v == "" or v is None:
                        col_data[h]["nulls"] += 1
                    if len(col_data[h]["values"]) < 200:
                        col_data[h]["values"].append(v)
        
        schema = {}
        for h in headers:
            d = col_data[h]
            null_pct = round(d["nulls"] / max(d["total"], 1) * 100, 2)
            uniq = len(set(str(v) for v in d["values"] if v is not None and str(v).strip()))
            dtype = infer_dtype(d["values"])
            schema[h] = {"dtype": dtype, "null_pct": null_pct, "unique_sample": uniq}
        
        return {
            "has_header": has_header,
            "columns": len(headers),
            "rows": row_count,
            "file_size": path.stat().st_size,
            "schema": schema,
        }
    except Exception as e:
        return {"error": str(e)}


def analyze_xlsx(path, max_rows=5000):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows = 0
            col_data = {}
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                    col_data = {h: {"nulls": 0, "total": 0, "values": []} for h in headers}
                else:
                    rows += 1
                    if rows > max_rows:
                        break
                    for j, h in enumerate(headers):
                        v = row[j] if j < len(row) else None
                        col_data[h]["total"] += 1
                        if v is None or (isinstance(v, str) and v.strip() == ""):
                            col_data[h]["nulls"] += 1
                        if len(col_data[h]["values"]) < 200:
                            col_data[h]["values"].append(v)
            schema = {}
            for h in headers:
                d = col_data.get(h, {"nulls": 0, "total": 1, "values": []})
                null_pct = round(d["nulls"] / max(d["total"], 1) * 100, 2)
                uniq = len(set(str(v) for v in d["values"] if v is not None and str(v).strip()))
                dtype = infer_dtype(d["values"])
                schema[h] = {"dtype": dtype, "null_pct": null_pct, "unique_sample": uniq}
            if rows > 0 or len(headers) > 1:
                result[sheet_name] = {
                    "columns": len(headers),
                    "rows": rows,
                    "schema": schema,
                }
        wb.close()
        if not result:
            return {"error": "No data sheets found"}
        return result
    except Exception as e:
        return {"error": str(e)}


def analyze_text(path, max_chars=2000):
    try:
        content = path.read_text("utf-8", errors="replace")
        lines = content.splitlines()
        return {
            "lines": len(lines),
            "size": path.stat().st_size,
            "preview": "\n".join(lines[:20])[:max_chars],
        }
    except Exception as e:
        return {"error": str(e)}


def analyze_zip(path):
    try:
        with zipfile.ZipFile(path, "r") as zf:
            names = zf.namelist()
            total_size = sum(zf.getinfo(n).file_size for n in names)
            return {
                "files": len(names),
                "total_uncompressed": total_size,
                "names": names[:50],
            }
    except Exception as e:
        return {"error": str(e)}


def analyze_file(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return analyze_csv(path)
    elif suffix in (".xlsx", ".xls"):
        return analyze_xlsx(path)
    elif suffix == ".zip":
        return analyze_zip(path)
    elif suffix == ".txt":
        return analyze_text(path)
    return {"format": suffix, "size": path.stat().st_size}


def main():
    print("=" * 80)
    print("PHASE 1: STUDY EVERY SOURCE — DEEP SCHEMA INFERENCE")
    print("=" * 80)
    
    sources = {
        "gdelt_events": DATASETS_DIR / "processed/gdelt/events/20240101/20240101000000.export.csv",
        "gdelt_mentions": DATASETS_DIR / "processed/gdelt/mentions/20240101/20240101000000.mentions.csv",
        "gdelt_gkg": DATASETS_DIR / "processed/gdelt/gkg/20240101/20240101000000.gkg.csv",
        "ofac_sanctions": DATASETS_DIR / "raw/sdn.csv",
        "ports": DATASETS_DIR / "raw/ports.csv",
        "global_energy_2025": DATASETS_DIR / "raw/global_energy_2025.csv",
        "global_energy_2026": DATASETS_DIR / "raw/global_energy_2026.csv",
    }
    
    gem_dir = DATASETS_DIR / "raw/gem-data"
    if gem_dir.exists():
        for f in sorted(gem_dir.iterdir()):
            if f.suffix.lower() in (".xlsx", ".xls"):
                name = f.stem.replace(" ", "_").replace(",", "").replace("(", "").replace(")", "")[:40]
                sources[f"gem_{name}"] = f
    
    for f in sorted(DATASETS_DIR.rglob("*")):
        if f.suffix.lower() in (".xlsx", ".xls") and f.parent.name == "raw":
            name = f.stem.replace(" ", "_").replace(",", "").replace("(", "").replace(")", "")[:40]
            sources[f"root_{name}"] = f
    
    aeo_txt = DATASETS_DIR / "raw/AEO2023/AEO2023.txt"
    if aeo_txt.exists():
        sources["aeo2023"] = aeo_txt
    aeo_txt2 = DATASETS_DIR / "raw/AEO2026/AEO2026.txt"
    if aeo_txt2.exists():
        sources["aeo2026"] = aeo_txt2
    
    results = {}
    
    for name, path in sorted(sources.items()):
        if not path.exists():
            continue
        
        print(f"\n--- Analyzing: {name} ({path.name}, {fmt_bytes(path.stat().st_size)}) ---")
        analysis = analyze_file(path)
        results[name] = analysis
        
        if "error" in analysis:
            print(f"  ERROR: {analysis['error']}")
            continue
        
        if isinstance(analysis, dict) and "schema" in analysis:
            s = analysis["schema"]
            cols = analysis.get("columns", len(s))
            rows = analysis.get("rows", "?")
            num_cols = sum(1 for v in s.values() if v.get("dtype") == "numeric")
            str_cols = sum(1 for v in s.values() if v.get("dtype") == "string")
            print(f"  {cols} cols, {rows} rows, {num_cols} numeric, {str_cols} string")
            for col, info in list(s.items())[:10]:
                print(f"    {col}: {info['dtype']} (nulls: {info['null_pct']}%, unique: {info['unique_sample']})")
            if len(s) > 10:
                print(f"    ... +{len(s)-10} more")
        
        elif isinstance(analysis, dict):
            sheets_analyzed = 0
            for sheet, s_data in analysis.items():
                if isinstance(s_data, dict) and "schema" in s_data and s_data.get("rows", 0) > 0:
                    sheets_analyzed += 1
                    cols = s_data.get("columns", 0)
                    rows = s_data.get("rows", 0)
                    if sheets_analyzed <= 3:
                        print(f"  Sheet [{sheet}]: {cols} cols, {rows} rows")
                        for col, info in list(s_data["schema"].items())[:5]:
                            print(f"    {col}: {info['dtype']} (nulls: {info['null_pct']}%)")
                        if len(s_data["schema"]) > 5:
                            print(f"    ... +{len(s_data['schema'])-5} more")
            if sheets_analyzed == 0:
                print(f"  No data sheets found")
    
    # Generate comprehensive report
    report = {
        "generated_at": datetime.utcnow().isoformat(),
        "sources": {},
    }
    
    for name, path in sorted(sources.items()):
        if not path.exists():
            continue
        
        analysis = results.get(name, {})
        rel_path = str(path.relative_to(DATASETS_DIR)) if DATASETS_DIR in path.parents else str(path)
        
        src_info = {
            "file": rel_path,
            "size": path.stat().st_size,
            "size_hr": fmt_bytes(path.stat().st_size),
            "analysis": analysis,
        }
        
        if isinstance(analysis, dict) and "error" in analysis:
            src_info["status"] = "error"
        else:
            src_info["status"] = "ok"
        
        if isinstance(analysis, dict) and "schema" in analysis:
            src_info["columns"] = analysis.get("columns", 0)
            src_info["rows"] = analysis.get("rows", 0)
            cols_list = []
            for col, info in analysis.get("schema", {}).items():
                cols_list.append({
                    "name": col,
                    "type": info.get("dtype", "unknown"),
                    "null_pct": info.get("null_pct", 0),
                    "unique_sample": info.get("unique_sample", 0),
                })
            src_info["columns_detail"] = cols_list
        
        report["sources"][name] = src_info
    
    report_path = OUTPUT_DIR / "source_study.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"\nDetailed report: {report_path}")
    
    # Generate data dictionary
    dict_path = OUTPUT_DIR / "DATA_DICTIONARY.md"
    with open(dict_path, "w", encoding="utf-8") as f:
        f.write("# Data Dictionary — Complete Source Catalog\n\n")
        f.write(f"**Generated:** {datetime.utcnow().isoformat()}\n\n")
        f.write("## Source Summary\n\n")
        f.write("| Source | File | Size | Cols | Rows | Status |\n")
        f.write("|---|---|---|---|---|---|\n")
        
        for name in sorted(report["sources"]):
            src = report["sources"][name]
            cols = src.get("columns", src.get("analysis", {}).get("columns", "?"))
            rows = src.get("rows", src.get("analysis", {}).get("rows", "?"))
            f.write(f"| {name} | {src['file']} | {src['size_hr']} | {cols} | {rows} | {src['status']} |\n")
        
        f.write("\n---\n\n")
        
        for name, src in sorted(report["sources"].items()):
            f.write(f"\n## {name}\n\n")
            f.write(f"- **File:** `{src['file']}`\n")
            f.write(f"- **Size:** {src['size_hr']}\n")
            f.write(f"- **Status:** {src['status']}\n")
            
            analysis = src.get("analysis", {})
            if "error" in analysis:
                f.write(f"- **Error:** {analysis['error']}\n\n")
                continue
            
            cols = src.get("columns", analysis.get("columns", 0))
            rows = src.get("rows", analysis.get("rows", 0))
            f.write(f"- **Columns:** {cols}\n")
            f.write(f"- **Rows:** {rows}\n")
            
            if "columns_detail" in src:
                f.write("\n### Schema\n\n")
                f.write("| # | Column | Type | Null % | Unique |\n")
                f.write("|---|---|---|---|---|\n")
                for i, col in enumerate(src["columns_detail"]):
                    f.write(f"| {i+1} | {col['name']} | {col['type']} | {col['null_pct']}% | {col['unique_sample']} |\n")
            
            f.write("\n---\n")
    
    print(f"Data dictionary: {dict_path}")
    print("Done.")


if __name__ == "__main__":
    main()
